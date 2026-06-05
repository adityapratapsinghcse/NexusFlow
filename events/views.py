import openpyxl,json
from django.db import models
from datetime import time
from datetime import datetime, timedelta, date
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from accounts.decorators import role_required
from django.contrib.auth.decorators import login_required
from accounts.models import COLLEGES, Profile
from notifications.utils import create_notification
from .forms import BulkCertificateForm, IndividualCertificateForm
from .models import Event, EventRegistration, StudentCertificate, Venue, WaitlistEntry ,EventFeedback,EventLogisticsAssignment
from notifications.models import Notification
from .utils import generate_qr_code, generate_unique_pass
from .validators import validate_document, validate_image
from django.core.exceptions import ValidationError
from accounts.models import log_action
from .email_utils import send_registration_email, send_pass_active_email, send_cancellation_email, send_event_rejection_email, send_event_approval_email
from .utils import promote_from_waitlist
from .agents.venue_agent import run_venue_agent
from .agents.review_agent import run_review_agent
from .agents.crowd_agent import run_crowd_agent

# ===========================================================
#  SCHEDULE EVENT  (Teacher / Admin)
#  This is the main view that now uses the AI agent.
# ===========================================================

@login_required
def schedule_event(request):
    venues = Venue.objects.filter(is_active=True).order_by('capacity')
    is_admin = request.user.groups.filter(name="Admin").exists()

    context = {
        "venues": venues,
        "colleges": Profile.objects.all(),
        "is_admin": is_admin,
    }

    if request.method == "POST":
        try:
            title = request.POST.get('title')
            description = request.POST.get('description')
            expected_crowd = int(request.POST.get('expected_crowd', 0))

            date_str = request.POST.get('date')
            start_str = request.POST.get('start_time')
            end_str = request.POST.get('end_time')

            registration_start_str = request.POST.get("registration_start")
            registration_end_str = request.POST.get("registration_end")

            poster = request.FILES.get('poster')
            override_venue_id = request.POST.get("override_venue")

            # ---- Convert date/time ----
            event_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            start_time_obj = datetime.strptime(start_str, "%H:%M").time()
            end_time_obj = datetime.strptime(end_str, "%H:%M").time()

            registration_start = timezone.make_aware(
                datetime.strptime(registration_start_str, "%Y-%m-%dT%H:%M")
            )
            registration_end = timezone.make_aware(
                datetime.strptime(registration_end_str, "%Y-%m-%dT%H:%M")
            )

        except (ValueError, TypeError, Exception):
            messages.error(request, "Invalid date/time context or input structural format.")
            return render(request, "schedule.html", context)

        # ---- Combine datetimes safely ----
        start_dt = timezone.make_aware(datetime.combine(event_date, start_time_obj))
        end_dt = timezone.make_aware(datetime.combine(event_date, end_time_obj))

        # ---- Basic Data Integrity Validations ----
        if end_dt <= start_dt:
            messages.error(request, "Operational end time must occur strictly after start execution.")
            return render(request, "schedule.html", context)

        if event_date < timezone.now().date():
            messages.error(request, "Target event parameter profile dates cannot look into the past.")
            return render(request, "schedule.html", context)

        if registration_end <= registration_start:
            messages.error(request, "Registration gate window close must be after opening timeline initialization.")
            return render(request, "schedule.html", context)

        # ---------------- AI CROWD ENGINE ANALYSIS ----------------
        try:
            crowd_result = run_crowd_agent(
                event_title=title,
                event_description=description
            )
            predicted_crowd = crowd_result.get('predicted_crowd', 0)
            ai_crowd_reason = crowd_result.get('reason', '')
            crowd_confidence = crowd_result.get('confidence', 'N/A')
        except Exception:
            predicted_crowd = expected_crowd
            ai_crowd_reason = "AI Prediction Node unreachable."
            crowd_confidence = "Low"

        # Dynamically scale infrastructure demands up if teacher parameters under-assessed scope
        if expected_crowd < predicted_crowd:
            expected_crowd = predicted_crowd

        # ---- Infrastructure Venue Distribution Processing ----
        selected_venue = None
        buffer = timedelta(hours=1)
        check_start_dt = start_dt - buffer
        check_end_dt = end_dt + buffer

        # 1️⃣ Admin Manual Bypass Override
        if is_admin and override_venue_id:
            override_venue = get_object_or_404(Venue, id=override_venue_id, is_active=True)
            clash = Event.objects.filter(
                venue=override_venue,
                start_datetime__lt=check_end_dt,
                end_datetime__gt=check_start_dt
            ).exists()

            if clash:
                messages.error(request, "The specified override venue allocation is blocked during these parameters.")
                return render(request, "schedule.html", context)
            selected_venue = override_venue

        # 2️⃣ AI Automation Agent Routine Routing
        if not selected_venue:
            is_night = (end_time_obj.hour >= 18)
            try:
                agent_result = run_venue_agent(
                    event_name=title,
                    event_description=description,
                    expected_crowd=expected_crowd,
                    start_dt_iso=start_dt.isoformat(),
                    end_dt_iso=end_dt.isoformat(),
                    is_night_event=is_night,
                )
                
                if not agent_result or not agent_result.get('venue_id'):
                    messages.error(request, "AI Scheduler Agent: No available allocation layouts could mitigate scheduling conflict clusters.")
                    return render(request, "schedule.html", context)

                selected_venue = Venue.objects.get(id=agent_result['venue_id'])
                ai_reason = agent_result.get('reason', 'Automated heuristic profiling allocation logic.')
                ai_warnings = agent_result.get('warnings', [])
            except Exception as e:
                messages.error(request, f"AI Allocation system module experienced exceptions: {str(e)}")
                return render(request, "schedule.html", context)
        else:
            ai_reason = "Manually locked override assigned via Admin Authorization Interface."
            ai_warnings = []

        # ---- User Identity Role Context Mapping ----
        if is_admin:
            initial_status = "APPROVED"
            is_general = request.POST.get("is_general") == "on"

            if is_general:
                college_profile = request.user.profile
            else:
                college_id = request.POST.get("college_profile")
                college_profile = get_object_or_404(Profile, id=college_id)
        elif request.user.groups.filter(name="Teacher").exists():
            initial_status = "PENDING"
            redirect_to = "teacher_dashboard"
            college_profile = request.user.profile
            is_general = False
        else:
            messages.error(request, "System Authorization Privileges evaluation failed. Access Denied.")
            return render(request, "schedule.html", context)

        # ---- Persistent Data Compilation & Instantiation ----
        event = Event.objects.create(
            organizer=request.user,
            college_profile=college_profile,
            venue=selected_venue,
            title=title,
            description=description,
            date=event_date,
            start_time=start_time_obj,
            end_time=end_time_obj,
            start_datetime=start_dt,
            end_datetime=end_dt,
            expected_crowd=expected_crowd,
            night_event_warning=(end_time_obj >= time(18, 0)),
            poster=poster,
            status=initial_status,
            is_general=is_general,
            registration_start=registration_start,
            registration_end=registration_end,
        )

        # Store prediction evaluation metrics safely
        event.predicted_crowd = predicted_crowd 
        event.crowd_confidence = crowd_confidence 
        event.ai_crowd_reason = ai_crowd_reason 
        event.ai_venue_reason = ai_reason
        event.ai_warnings = json.dumps(ai_warnings) if ai_warnings else None
        event.save()

        # ---- Broadcast Lifecycle Notifications ----
        Notification.objects.create(
            role="ADMIN",
            title="New Operational Core Registration Event",
            message=f"{request.user.get_full_name()} created/proposed operational context execution '{title}'"
        )
        # ---------------- AI RESULT CONTEXT ---------------- 
        context['ai_prediction'] = { 
            "predicted_crowd": predicted_crowd, 
            "confidence": crowd_confidence, 
            "reason": ai_crowd_reason, 
            "venue_reason": ai_reason, 
            "venue_name": selected_venue.name, 
            "warnings": ai_warnings, 
        }
        messages.success( request, f"Event scheduled successfully in {selected_venue.name}" ) 
        return render(request, "schedule.html", context)
    return render(request, "schedule.html", context)
#=========================================================================================================

# ===========================================================
#  ACTION EVENT (Admin approve/reject list)
# ===========================================================

@login_required
@role_required(['Admin'])
def action_event(request):
    pending_events = Event.objects.filter(
        status='PENDING'
    ).order_by('-created_at').select_related('venue', 'organizer')

    # Attach AI review to each event
    for event in pending_events:
        try:
            review = run_review_agent(event)
            event.ai_review = review  # attached as attribute, not saved
        except Exception:
            event.ai_review = {
                'summary': 'AI review unavailable.',
                'recommendation': 'APPROVE WITH CAUTION',
                'risk_flags': []
            }

    return render(request, "action.html", {'events': pending_events})

#=========================================================================================================

# ===========================================================
#  UPDATE EVENT STATUS (Admin approve/reject)
# ===========================================================
@login_required
@role_required(['Admin'])
def update_event(request, event_id, action):
    event = get_object_or_404(Event, id=event_id)

    if action == 'approve':
        event.status = 'APPROVED'
        event.save()
        send_event_approval_email(event.organizer, event)
        log_action(request.user, 'EVENT_APPROVED', event=event, request=request)
        create_notification(role='TEACHER', user=event.organizer, title='Event Approved',
                            message=f"Your event '{event.title}' has been approved.")
        create_notification(role='STUDENT', title='New Event Live',
                            message=f"'{event.title}' is now open for registration.")
        messages.success(request, f"Event '{event.title}' approved.")

    elif action == 'reject':
        reason = request.POST.get('rejection_reason', '').strip()
        if not reason:
            messages.error(request, "Please provide a rejection reason.")
            return redirect('action_event')
        event.status = 'REJECTED'
        event.rejection_reason = reason
        event.save()
        send_event_rejection_email(event.organizer, event, reason)
        log_action(request.user, 'EVENT_REJECTED', event=event, notes=reason, request=request)
        create_notification(
            role='TEACHER', user=event.organizer,
            title='Event Rejected',
            message=f"Your event '{event.title}' was rejected. Reason: {reason}",
            event=event,
        )
        messages.warning(request, f"Event '{event.title}' rejected.")
    
    if action not in ('approve', 'reject'):
        messages.error(request, "Invalid action.")
        return redirect('action_event')

    return redirect('action_event')




# ===========================================================
#  EVENT DETAIL + REGISTRATION (Student)
# ===========================================================

@login_required
def event_detail(request, event_id):
    event = get_object_or_404(Event, id=event_id, status="APPROVED")

    already_registered = EventRegistration.objects.filter(
        student=request.user, event=event
    ).exists()

    if request.method == "POST":
        now_time = timezone.now()
        if event.registration_start and now_time < event.registration_start:
            messages.error(request, f"Registration opens on {event.registration_start.strftime('%d %b %Y at %I:%M %p')}.")
            return redirect("student_events")

        if event.registration_end and now_time > event.registration_end:
            messages.error(request, "Registration for this event has closed.")
            return redirect("student_events")
        if already_registered:
            messages.info(request, "You are already registered for this event.")

        elif hasattr(request.user, 'profile') and request.user.profile.is_banned:
            messages.error(request, "You are banned from registering for events.")
            return redirect("student_events")
        else:
            current_count = EventRegistration.objects.filter(event=event).count()
            if current_count >= event.expected_crowd:
                already_waiting = WaitlistEntry.objects.filter(
                    student=request.user, event=event
                ).exists()
                if not already_waiting:
                    WaitlistEntry.objects.create(student=request.user, event=event)
                    messages.info(request, "Event is full. You've been added to the waitlist.")
                else:
                    messages.info(request, "You're already on the waitlist for this event.")
                return redirect("student_events")

            reg = EventRegistration.objects.create(
                student=request.user, event=event,
                verified=False, pass_active=False
            )
            generate_unique_pass(reg)
            generate_qr_code(reg)
            send_registration_email(request.user, event)
            create_notification(role='TEACHER', user=event.organizer,
                                title='New Student Registration',
                                message=f"{request.user.get_full_name()} registered for '{event.title}'.")

            if current_count + 1 >= event.expected_crowd:
                create_notification(role='TEACHER', user=event.organizer,
                                    title='Event Full',
                                    message=f"'{event.title}' has reached its expected crowd size.")

            messages.success(request, "Successfully registered! Your pass will be issued soon.")
        return redirect("student_events")

    return render(request, "student_register.html", {
        "event":              event,
        "already_registered": already_registered,
    })


# ===========================================================
#  MY REGISTRATIONS (Student)
# ===========================================================

@login_required
def my_registrations(request):
    registrations = EventRegistration.objects.filter(
        student=request.user
    ).select_related('event', 'event__venue').order_by('event__date', 'event__start_time')
    return render(request, "my_registrations.html", {"registrations": registrations})


@login_required
def cancel_registration(request, event_id):
    registration = get_object_or_404(
        EventRegistration, event_id=event_id, student=request.user
    )
    event = registration.event

    # Block if pass already activated
    if registration.pass_active:
        messages.error(request, "Cannot cancel — your pass is already active.")
        return redirect('my_registrations')

    # Block within 24 hours of event start
    if event.start_datetime and timezone.now() > event.start_datetime - timedelta(hours=24):
        messages.error(request, "Cannot cancel within 24 hours of the event.")
        return redirect('my_registrations')

    if request.method == 'POST':
        registration.delete()
        promote_from_waitlist(registration.event)
        log_action(request.user, 'REG_CANCELLED', event=event, request=request)
        create_notification(
            role='TEACHER',
            user=event.organizer,
            title='Registration Cancelled',
            message=f"{request.user.get_full_name()} cancelled their registration for '{event.title}'.",
            event=event,
        )
        messages.success(request, f"Registration for '{event.title}' cancelled.")
        return redirect('my_registrations')

    return render(request, 'cancel_registration_confirm.html', {'event': event})

# ===========================================================
#  UPCOMING EVENTS (Public browse)
# ===========================================================

@login_required
def upcoming_events(request):
    today = date.today()
    search_query = request.GET.get('search', '').strip()
    venue_filter = request.GET.get('venue', '')
    date_filter = request.GET.get('date_sort', 'soonest')

    events = Event.objects.filter(status="APPROVED", date__gte=today).select_related('venue')
    if search_query:
        events = events.filter(Q(title__icontains=search_query) | Q(description__icontains=search_query))
    if venue_filter:
        events = events.filter(venue__id=venue_filter)

    if date_filter == 'soonest':
        events = events.order_by("date", "start_time")
    else:
        events = events.order_by("-date")
    venues = Venue.objects.filter(is_active=True)
    registered_event_ids = list(EventRegistration.objects.filter(
        student=request.user
    ).values_list("event_id", flat=True))

    is_student = request.user.groups.filter(name='Student').exists()

    return render(request, "upcoming_events.html", {
        "events": events,
        "venues": venues,
        "search_query": search_query,
        "selected_venue": venue_filter,
        "selected_sort": date_filter,
        "registered_event_ids": registered_event_ids, # Sent to template layout engine
        "is_student" : is_student,
    })


# ===========================================================
#  CHECK STATUS (Teacher sees their own events)
# ===========================================================

@login_required
def check_status(request):
    today          = date.today()
    next_two_months = today + timedelta(days=60)
    events_status  = Event.objects.filter(
        organizer=request.user, date__range=[today, next_two_months]
    ).order_by("date").select_related('venue')
    return render(request, "check_status.html", {"events": events_status})


# ===========================================================
#  CANCEL EVENT
# ===========================================================

@login_required
def cancel_event(request, event_id):
    user = request.user
    if user.groups.filter(name="Admin").exists():
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user, status="PENDING")

    if request.method == "POST":
        event_name = event.title
        for reg in EventRegistration.objects.filter(event=event).select_related('student'):
            send_cancellation_email(reg.student, event)
        event.delete()
        messages.success(request, f"Event '{event_name}' cancelled successfully.")
        return redirect("cancel_event_list")

    return render(request, "cancel_event_confirm.html", {"event": event})


@login_required
def cancel_event_list(request):
    user = request.user
    search_query = request.GET.get('search', '')
    venue_filter = request.GET.get('venue', '')
    date_filter  = request.GET.get('date_sort', 'soonest')

    if user.groups.filter(name="Admin").exists():
        events = Event.objects.all().order_by("-date")
    else:
        events = Event.objects.filter(organizer=user, status="PENDING").order_by("date")

    if search_query:
        events = events.filter(Q(title__icontains=search_query) | Q(description__icontains=search_query))
    if venue_filter:
        events = events.filter(venue__id=venue_filter)

    events = events.order_by("date", "start_time") if date_filter == 'soonest' else events.order_by("-date")
    venues = Venue.objects.filter(is_active=True)

    return render(request, "cancel_event.html", {
        "events":         events,
        "venues":         venues,
        "search_query":   search_query,
        "selected_venue": venue_filter,
        "selected_sort":  date_filter,
    })


# ===========================================================
#  TEACHERS EVENT LIST
# ===========================================================

@login_required
def teachers_event_list(request):
    today        = date.today()
    search_query = request.GET.get('search', '')
    venue_filter = request.GET.get('venue', '')
    date_filter  = request.GET.get('date_sort', 'soonest')
    user         = request.user

    if user.groups.filter(name="Admin").exists():
        events = Event.objects.all().order_by("-date")
    else:
        events = Event.objects.filter(
            organizer=user, status__in=['PENDING', 'APPROVED'], date__gte=today
        ).order_by("date")

    if search_query:
        events = events.filter(Q(title__icontains=search_query) | Q(description__icontains=search_query))
    if venue_filter:
        events = events.filter(venue__id=venue_filter)

    events = events.order_by("date", "start_time") if date_filter == 'soonest' else events.order_by("-date")
    venues = Venue.objects.filter(is_active=True)

    return render(request, "teacher_event_list.html", {
        "events":         events,
        "venues":         venues,
        "search_query":   search_query,
        "selected_venue": venue_filter,
        "selected_sort":  date_filter,
    })

@login_required
def reassign_venue(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    user  = request.user
    is_admin = user.groups.filter(name='Admin').exists()

    # Permission: admin can always; teacher only for their own PENDING event
    if not is_admin:
        if event.organizer != user:
            messages.error(request, "You can only reassign your own events.")
            return redirect('teacher_my_events')
        if event.status != 'PENDING':
            messages.error(request, "Venue can only be changed while event is PENDING. Contact admin for approved events.")
            return redirect('check_status')
        # Teachers can only override within 72h of creation
        from datetime import timedelta
        if timezone.now() > event.created_at + timedelta(hours=72):
            messages.error(request, "72-hour self-reassignment window has expired. Contact admin.")
            return redirect('check_status')

    venues = Venue.objects.filter(
        is_active=True,
        capacity__gte=event.expected_crowd
    ).order_by('capacity')

    if request.method == 'POST':
        venue_id = request.POST.get('venue_id')
        reason   = request.POST.get('reason', '').strip()

        if not reason or len(reason) < 20:
            messages.error(request, "Please provide a detailed reason (at least 20 characters).")
            return render(request, 'reassign_venue.html', {'event': event, 'venues': venues, 'is_admin': is_admin})

        new_venue = get_object_or_404(Venue, id=venue_id, is_active=True)

        # Capacity check
        if new_venue.capacity < event.expected_crowd:
            messages.error(request, f"'{new_venue.name}' capacity ({new_venue.capacity}) is less than expected crowd ({event.expected_crowd}).")
            return render(request, 'reassign_venue.html', {'event': event, 'venues': venues, 'is_admin': is_admin})

        # Clash check
        buffer = timedelta(hours=1)
        clash = Event.objects.filter(
            venue=new_venue,
            start_datetime__lt=event.end_datetime + buffer,
            end_datetime__gt=event.start_datetime - buffer,
        ).exclude(id=event.id).exists()

        if clash:
            messages.error(request, f"'{new_venue.name}' is already booked during this time.")
            return render(request, 'reassign_venue.html', {'event': event, 'venues': venues, 'is_admin': is_admin})

        old_venue_name = event.venue.name if event.venue else '(none)'
        event.venue = new_venue
        event.save(update_fields=['venue'])

        log_action(
            user, 'VENUE_OVERRIDE', event=event,
            notes=f"{'Admin' if is_admin else 'Teacher'} changed venue from '{old_venue_name}' to '{new_venue.name}'. Reason: {reason}",
            request=request,
        )

        # If teacher override, notify admin
        if not is_admin:
            create_notification(
                role='ADMIN',
                title='Venue Changed by Teacher',
                message=f"{user.get_full_name()} changed venue for '{event.title}' from '{old_venue_name}' to '{new_venue.name}'. Reason: {reason}",
                event=event,
            )

        messages.success(request, f"Venue changed to '{new_venue.name}'.")
        return redirect('teacher_my_events' if not is_admin else 'teachers_event_list')

    return render(request, 'reassign_venue.html', {
        'event': event, 'venues': venues, 'is_admin': is_admin,
        'can_override': True,
    })

# ===========================================================
#  UPDATE EVENT (Teacher / Admin edit)
# ===========================================================

@login_required
def update_event_teacher(request, event_id):
    user  = request.user
    event = get_object_or_404(Event, id=event_id)

    if user.groups.filter(name="Teacher").exists() and event.organizer != user:
        messages.error(request, "You are not allowed to edit this event.")
        return redirect("main_page")
    elif not user.groups.filter(name__in=["Teacher", "Admin"]).exists():
        messages.error(request, "Unauthorised access.")
        return redirect("main_page")

    if request.method == "POST":
        text_changed   = False
        poster_changed = False
        text_fields    = ['title', 'description', 'date', 'start_time', 'end_time',
                          'expected_crowd', 'registration_start', 'registration_end']

        for field in text_fields:
            new_val = request.POST.get(field)
            old_val = getattr(event, field)
            if not new_val:
                continue
            if field == 'date':
                new_val = datetime.strptime(new_val, '%Y-%m-%d').date()
            elif field in ['start_time', 'end_time']:
                new_val = datetime.strptime(new_val, '%H:%M').time()
            elif field in ['registration_start', 'registration_end']:
                new_val = timezone.make_aware(datetime.strptime(new_val, '%Y-%m-%dT%H:%M'))
            elif field == 'expected_crowd':
                new_val = int(new_val)
            if new_val != old_val:
                setattr(event, field, new_val)
                text_changed = True

        new_poster = request.FILES.get('poster')
        if new_poster:
            event.poster  = new_poster
            poster_changed = True

        if user.groups.filter(name="Teacher").exists() and text_changed and event.status != "PENDING":
            event.status = "PENDING"
        elif user.groups.filter(name="Admin").exists() and event.status != "APPROVED":
            event.status = "APPROVED"

        if text_changed or poster_changed:
            event.save()
            if text_changed and user.groups.filter(name="Teacher").exists():
                messages.success(request, "Event updated and resent for approval.")
            elif poster_changed and not text_changed:
                messages.info(request, "Poster updated. Status unchanged.")
            else:
                messages.success(request, "Event updated successfully.")
        else:
            messages.info(request, "No changes detected.")

        return redirect('teachers_event_list')

    return render(request, "update_event.html", {"events": event})


# ===========================================================
#  STUDENTS REGISTERED (Teacher/Admin view)
# ===========================================================

@login_required
def students_registered(request):
    user = request.user
    if user.groups.filter(name='Admin').exists():
        events = Event.objects.all().annotate(student_count=Count('eventregistration'))
    else:
        events = Event.objects.filter(organizer=user,status="APPROVED").annotate(student_count=Count('eventregistration'))
    context = {
        "events": events,
        "now": timezone.now() 
    }
    return render(request, "students_registered.html", context)


# ===========================================================
#  EVENT STUDENT DETAILS
# ===========================================================

@login_required
def event_student_details(request, event_id):
    user = request.user
    if user.groups.filter(name='Admin').exists():
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)

    students = EventRegistration.objects.filter(event=event).select_related('student')
    return render(request, "event_students_detail.html", {"event": event, "students": students})


# ===========================================================
#  EXPORT TO EXCEL
# ===========================================================

@login_required
def export_event_students(request, event_id):
    """Old all-registrations export — kept for backwards compat."""
    user = request.user
    if user.groups.filter(name='Admin').exists():
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)

    registrations = EventRegistration.objects.filter(event=event).select_related('student')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"
    ws.append(['#', 'Name', 'Email', 'Pass Code', 'Registered At', 'Attended'])

    for idx, reg in enumerate(registrations, start=1):
        ws.append([
            idx,
            reg.student.get_full_name(),
            reg.student.email,
            reg.unique_pass or "-",
            reg.registered_at.strftime("%Y-%m-%d %H:%M"),
            "Yes" if reg.verified else "No",
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=registrations_{event.title}.xlsx'
    wb.save(response)
    return response


@login_required
def export_attendance_report(request, event_id):
    """
    Full branded attendance report export — all fields from Event,
    EventRegistration and Profile / CollegeData for every participant.
    Downloaded as  Attendance_Report_<EventTitle>.xlsx
    """
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    user = request.user
    is_admin = user.groups.filter(name='Admin').exists()
    is_assigned = EventLogisticsAssignment.objects.filter(
        event_id=event_id, assigned_teacher=user
    ).exists()

    if is_admin:
        event = get_object_or_404(Event, id=event_id)
    elif is_assigned:
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)

    registrations = (
        EventRegistration.objects
        .filter(event=event)
        .select_related('student', 'student__profile', 'student__profile__college_data')
        .order_by('-verified', 'student__last_name')
    )

    # ── Workbook setup ────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Theme colours matching NexusFlow dark theme
    DARK_BG   = "060913"
    CARD_BG   = "0B1120"
    CYAN      = "00F0FF"
    GREEN     = "39FF14"
    YELLOW    = "FFAA00"
    WHITE     = "FFFFFF"
    MUTED     = "64748B"
    LIGHT     = "CBD5E1"
    HEADER_BG = "0F1A2E"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(hex_color=WHITE, bold=False, size=10):
        return Font(color=hex_color, bold=bold, size=size, name="Calibri")

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="1E293B")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title block (rows 1-4) ────────────────────────────────────────
    ws.merge_cells("A1:J1")
    ws["A1"] = "NEXUSFLOW — FINAL PARTICIPATION REPORT"
    ws["A1"].font = Font(color=CYAN, bold=True, size=14, name="Calibri")
    ws["A1"].fill = fill(DARK_BG)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Event: {event.title}"
    ws["A2"].font = Font(color=WHITE, bold=True, size=11, name="Calibri")
    ws["A2"].fill = fill(CARD_BG)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    ws.merge_cells("A3:J3")
    venue_name = event.venue.name if event.venue else "N/A"
    organizer_name = event.organizer.get_full_name() if event.organizer else "N/A"
    ws["A3"] = (
        f"Date: {event.date.strftime('%d %B %Y')}   |   "
        f"Time: {event.start_time.strftime('%I:%M %p')} - {event.end_time.strftime('%I:%M %p')}   |   "
        f"Venue: {venue_name}   |   "
        f"Organiser: {organizer_name}   |   "
        f"Capacity: {event.expected_crowd}"
    )
    ws["A3"].font = Font(color=MUTED, size=9, name="Calibri")
    ws["A3"].fill = fill(CARD_BG)
    ws["A3"].alignment = center()
    ws.row_dimensions[3].height = 16

    # Stats row (row 4)
    total_reg    = registrations.count()
    total_verif  = registrations.filter(verified=True).count()
    total_absent = total_reg - total_verif

    ws.merge_cells("A4:C4")
    ws["A4"] = f"Total Registered: {total_reg}"
    ws["A4"].font = Font(color=CYAN, bold=True, size=10, name="Calibri")
    ws["A4"].fill = fill(HEADER_BG)
    ws["A4"].alignment = center()

    ws.merge_cells("D4:F4")
    ws["D4"] = f"Verified Entry: {total_verif}"
    ws["D4"].font = Font(color=GREEN, bold=True, size=10, name="Calibri")
    ws["D4"].fill = fill(HEADER_BG)
    ws["D4"].alignment = center()

    ws.merge_cells("G4:J4")
    ws["G4"] = f"Not Scanned: {total_absent}"
    ws["G4"].font = Font(color=YELLOW, bold=True, size=10, name="Calibri")
    ws["G4"].fill = fill(HEADER_BG)
    ws["G4"].alignment = center()

    ws.row_dimensions[4].height = 18
    ws.append([])  # spacer row 5

    # ── Column headers (row 6) ────────────────────────────────────────
    HEADERS = [
        "#",
        "Full Name",
        "Student ID",
        "Email Address",
        "College",
        "Branch / Dept",
        "Phone",
        "Pass Code",
        "Arrival Time",
        "Status",
    ]
    ws.append(HEADERS)
    header_row = ws.max_row

    for col_idx, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill    = fill(HEADER_BG)
        cell.font    = Font(color=CYAN, bold=True, size=9, name="Calibri")
        cell.alignment = center()
        cell.border  = border
    ws.row_dimensions[header_row].height = 20

    # ── Data rows ─────────────────────────────────────────────────────
    COLLEGE_MAP = dict(COLLEGES)

    for idx, reg in enumerate(registrations, start=1):
        student = reg.student
        try:
            profile     = student.profile
            cd          = profile.college_data
            college_str = COLLEGE_MAP.get(profile.college_name, profile.college_name)
            branch_str  = cd.branch or "—" if cd else "—"
            phone_str   = cd.phone  or profile.phone or "—" if cd else (profile.phone or "—")
        except Exception:
            college_str = "—"
            branch_str  = "—"
            phone_str   = "—"

        arrival = (
            reg.verified_at.strftime("%d %b %Y %I:%M %p")
            if reg.verified_at else "—"
        )
        status = "VERIFIED" if reg.verified else "NOT SCANNED"

        row_data = [
            idx,
            student.get_full_name(),
            student.username,
            student.email,
            college_str,
            branch_str,
            phone_str,
            reg.unique_pass or "—",
            arrival,
            status,
        ]
        ws.append(row_data)
        data_row = ws.max_row
        ws.row_dimensions[data_row].height = 18

        # Row fill — alternate banding + verified highlight
        if reg.verified:
            row_bg = "0D1F0D"   # subtle green tint
            status_color = GREEN
        else:
            row_bg = CARD_BG if idx % 2 == 0 else "0D1426"
            status_color = YELLOW

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.fill    = fill(row_bg)
            cell.border  = border
            cell.alignment = center() if col_idx in (1, 9, 10) else left()

            if col_idx == 2:  # name — bold white
                cell.font = Font(color=WHITE, bold=True, size=10, name="Calibri")
            elif col_idx == 10:  # status
                cell.font = Font(color=status_color, bold=True, size=9, name="Calibri")
            elif col_idx == 8:  # pass code — monospace-ish
                cell.font = Font(color=CYAN, size=9, name="Courier New")
            elif col_idx == 9:  # arrival time
                cell.font = Font(color=CYAN if reg.verified else MUTED, size=9, name="Calibri")
            else:
                cell.font = Font(color=LIGHT, size=10, name="Calibri")

    # ── Column widths ─────────────────────────────────────────────────
    col_widths = [5, 26, 16, 30, 30, 20, 16, 14, 22, 14]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Freeze header rows so they stay visible while scrolling
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Download response ─────────────────────────────────────────────
    safe_title = "".join(c for c in event.title if c.isalnum() or c in (" ", "_", "-")).strip()
    filename = f"Attendance_Report_{safe_title}_{event.date.strftime('%Y%m%d')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ===========================================================
#  MY PASSES (Student)
# ===========================================================

@login_required
def my_passes(request):
    registrations = EventRegistration.objects.filter(
        student=request.user, pass_active=True
    ).select_related('event', 'event__venue')
    return render(request, "my_passes.html", {"registrations": registrations})


# ===========================================================
#  TEACHER MY EVENTS
# ===========================================================

@login_required
def teacher_events(request):
    today  = timezone.localdate()
    now    = timezone.now()
    events = Event.objects.filter(
        organizer=request.user, date__gte=today,status='APPROVED'
    ).order_by("date").select_related('venue')

    is_admin = request.user.groups.filter(name="Admin").exists()

    for event in events:
        event.can_publish  = not event.is_published
        event.total_registrations = event.eventregistration_set.count()
        event.can_reassign = (timezone.now() <= event.created_at + timedelta(hours=72))

    return render(request, "teacher_my_events.html", {
        "events":   events,
        "is_admin": is_admin,
        "today": now,
    })


# ============================================================
#  PUBLISH PASSES (FIXED — Admin + Assigned Teacher allowed)
# ============================================================
@login_required
def publish_passes(request, event_id):
    user = request.user
    is_admin = user.groups.filter(name='Admin').exists()

    # Check if user is organizer, admin, or assigned logistics teacher
    is_assigned = EventLogisticsAssignment.objects.filter(
        event_id=event_id,
        assigned_teacher=user
    ).exists()

    if is_admin:
        event = get_object_or_404(Event, id=event_id)
    elif is_assigned:
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)

    registrations = EventRegistration.objects.filter(event=event)
    if registrations.count() == 0:
        messages.warning(request, "No students are registered for this event. Passes cannot be published yet.")
        if is_admin:
            return redirect('admin_logistics_overview')
        return redirect('teacher_my_events')
    registrations.update(pass_active=True)
    event.is_published = True
    event.save(update_fields=['is_published'])

    action_type = 'PASS_PUBLISHED_ADMIN' if is_admin else 'PASS_PUBLISHED'
    log_action(user, action_type, event=event, request=request)

    for reg in registrations.select_related('student'):
        create_notification(
            role='STUDENT',
            user=reg.student,
            title='Your Event Pass is Active!',
            message=f"Your pass for '{event.title}' is now active. Check My Passes to get your QR code.",
            event=event,
        )
        send_pass_active_email(reg.student, event, reg.unique_pass)

    messages.success(request, f"Passes for '{event.title}' are now active!")
    if is_admin:
        return redirect('admin_logistics_overview')
    elif is_assigned:
        return redirect('teacher_logistics_dashboard')
    else:
        return redirect('teacher_my_events')

# ===========================================================
#  STUDENTS PARTICIPATED (Attendance list)
# ===========================================================

@login_required
def students_participated(request, event_id):
    user = request.user
    is_admin = user.groups.filter(name='Admin').exists()
    is_assigned = EventLogisticsAssignment.objects.filter(
        event_id=event_id, assigned_teacher=user
    ).exists()

    if is_admin or is_assigned:
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)
    
    registrations = EventRegistration.objects.filter(
        event=event
    ).select_related('student')
    return render(request, "students_participated.html", {
            "registrations": registrations,
            "event": event})

@login_required
@require_POST
def verify_pass(request, event_id):
    code = request.POST.get("code", "").strip().upper()

    if not code:
        return JsonResponse({"status": "invalid", "message": "No code received."})

    # Use select_related to fetch student name in the same query
    try:
        registration = EventRegistration.objects.select_related('student', 'event').get(
            event_id=event_id,
            unique_pass=code,
            pass_active=True,
        )
    except EventRegistration.DoesNotExist:
        return JsonResponse({
            "status":  "invalid",
            "message": "Pass not found or not yet activated.",
        })

    # If already verified, return a clear message (don't update again)
    if registration.verified:
        return JsonResponse({
            "status":       "already_used",
            "student_name": registration.student.get_full_name(),
            "verified_at":  registration.verified_at.strftime("%I:%M %p") if registration.verified_at else "earlier",
            "event_name":   registration.event.title,
        })

    # ATOMIC UPDATE — safe if two phones scan the exact same QR at the same time.
    # .filter(verified=False).update(...) is a single SQL UPDATE statement.
    # If another request beat us to it, `updated` will be 0 (zero rows updated).
    updated = EventRegistration.objects.filter(
        pk=registration.pk,
        verified=False         # Only update if STILL unverified
    ).update(
        verified=True,
        verified_at=timezone.now()
    )

    if updated == 1:
        # We successfully marked this student as verified
        return JsonResponse({
            "status":       "valid",
            "student_name": registration.student.get_full_name(),
            "event_name":   registration.event.title,
            "pass_code":    code,
        })
    else:
        # Race condition: another scan beat us by a millisecond
        return JsonResponse({
            "status":       "already_used",
            "student_name": registration.student.get_full_name(),
            "verified_at":  "just now",
            "event_name":   registration.event.title,
        })


# ===========================================================
#  SCAN PASS (Opens the scanner page)
# ===========================================================

@login_required
def scan_pass(request, event_id):
    user = request.user
    event = get_object_or_404(Event, id=event_id)

    is_admin = user.groups.filter(name='Admin').exists()
    is_organizer = event.organizer == user
    is_assigned = EventLogisticsAssignment.objects.filter(
        event=event, assigned_teacher=user
    ).exists()

    if not (is_admin or is_organizer or is_assigned):
        messages.error(request, "You don't have permission to scan passes for this event.")
        return redirect('teacher_dashboard')

    return render(request, 'scan_pass.html', {'event': event})


# ============================================================
#  VERIFY PASSES LIST (FIXED — shows assigned events too)
# ============================================================
@login_required
def verify_passes_list(request):
    user = request.user
    is_admin = user.groups.filter(name='Admin').exists()

    if is_admin:
        # Admin sees ALL approved events
        events = Event.objects.filter(
            status='APPROVED'
        ).order_by('-date').select_related('venue', 'organizer')
    else:
        # Teacher sees their own events + events they're assigned to
        own_events = Event.objects.filter(organizer=user)
        assigned_event_ids = EventLogisticsAssignment.objects.filter(
            assigned_teacher=user
        ).values_list('event_id', flat=True)
        assigned_events = Event.objects.filter(id__in=assigned_event_ids)

        # Combine and deduplicate
        from itertools import chain
        from django.db.models import Q
        events = Event.objects.filter(
            Q(organizer=user) | Q(id__in=assigned_event_ids),status="APPROVED"
        ).order_by('-date').select_related('venue')

    return render(request, 'verify_event_list.html', {'events': events})

# ===========================================================
#  MANAGE EVENTS
# ===========================================================

@login_required
def manage_events(request):
    events = Event.objects.filter(
        organizer=request.user,status = "APPROVED"
    ).order_by('-date').select_related('venue')
    return render(request, "manage_events.html", {"events": events})


# ===========================================================
#  UPLOAD BUS ROUTE
# ===========================================================

@login_required
def upload_bus_route(request, event_id):
    user = request.user
    is_admin = user.groups.filter(name='Admin').exists()
    is_assigned = EventLogisticsAssignment.objects.filter(
        event_id=event_id, assigned_teacher=user
    ).exists()

    if is_admin:
        event = get_object_or_404(Event, id=event_id)
    elif is_assigned:
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)

    if request.method == "POST":
        bus_route_file = request.FILES.get('bus_route_file')
        if bus_route_file:
            try:
                validate_document(bus_route_file)
            except ValidationError as e:
                messages.error(request, e.message)
                return render(request, "upload_bus_route.html", {"event": event})
            event.bus_route = bus_route_file
            event.save(update_fields=['bus_route'])
            if is_assigned:
                return redirect('teacher_logistics_dashboard')
            if is_admin:
                return redirect('admin_logistics_overview')
        else:
            messages.error(request, "Please select a file to upload.")

    return render(request, "upload_bus_route.html", {"event": event})


# ===========================================================
#  DISPLAY BUS ROUTES (Student)
# ===========================================================

@login_required
def display_bus_routes(request):
    events_with_routes = []
    registrations = EventRegistration.objects.filter(
        student=request.user
    ).select_related('event').order_by('event__date')

    for reg in registrations:
        if reg.event.bus_route:
            events_with_routes.append(reg.event)

    return render(request, "display_bus_routes.html", {"events": events_with_routes})


# ===========================================================
#  UPLOAD CERTIFICATE (Teacher issues certificates)
# ===========================================================

@login_required
def upload_certificate(request, event_id):
    user = request.user
    is_admin = user.groups.filter(name='Admin').exists()
    is_assigned = EventLogisticsAssignment.objects.filter(
        event_id=event_id, assigned_teacher=user
    ).exists()
    if is_admin or is_assigned:
        event = get_object_or_404(Event, id=event_id)
    else:
        event = get_object_or_404(Event, id=event_id, organizer=user)

    bulk_form       = BulkCertificateForm()
    individual_form = IndividualCertificateForm()
    students_registered = EventRegistration.objects.filter(event=event).select_related('student')
    bulk_form.fields['selected_students_bulk'].queryset = students_registered

    if request.method == "POST":
        if 'bulk_file' in request.FILES:
            bulk_form = BulkCertificateForm(request.POST, request.FILES)
            bulk_form.fields['selected_students_bulk'].queryset = students_registered
            if bulk_form.is_valid():
                file         = bulk_form.cleaned_data['bulk_file']
                registrations = bulk_form.cleaned_data['selected_students_bulk'] or students_registered
                for reg in registrations:
                    StudentCertificate.objects.create(
                        student=reg.student, event=event, certificate_file=file
                    )
                messages.success(request, f"Certificates uploaded for {registrations.count()} students.")
                if request.user.groups.filter(name='Admin').exists():
                    return redirect('teachers_event_list')
                return redirect('manage_events')

        elif 'certificate_file' in request.FILES:
            individual_form = IndividualCertificateForm(request.POST, request.FILES)
            if individual_form.is_valid():
                username         = individual_form.cleaned_data['student_username']
                certificate_file = individual_form.cleaned_data['certificate_file']
                student_photo    = individual_form.cleaned_data.get('student_photo')
                try:
                    student = User.objects.get(username=username)
                    if not EventRegistration.objects.filter(event=event, student=student).exists():
                        messages.error(request, f"{student.get_full_name()} did not participate in this event.")
                        return redirect('upload_certificate', event_id=event.id)
                except User.DoesNotExist:
                    messages.error(request, f"User '{username}' does not exist.")
                    return redirect('upload_certificate', event_id=event.id)

                StudentCertificate.objects.create(
                    student=student, event=event,
                    certificate_file=certificate_file, photo=student_photo
                )
                messages.success(request, f"Certificate uploaded for {student.get_full_name()}.")
                if request.user.groups.filter(name='Admin').exists():
                    return redirect('teachers_event_list')
                is_assigned_user = EventLogisticsAssignment.objects.filter(
                    event_id=event_id, assigned_teacher=request.user
                ).exists()
                if is_assigned_user:
                    return redirect('teacher_logistics_dashboard')
                return redirect('manage_events')

    return render(request, "upload_certificate.html", {
        "event":               event,
        "bulk_form":           bulk_form,
        "individual_form":     individual_form,
        "students_registered": students_registered,
    })


# ===========================================================
#  AJAX: Student autocomplete for certificate upload
# ===========================================================

@login_required
def ajax_get_event_students(request, event_id):
    query = request.GET.get('q', '')
    event = get_object_or_404(Event, id=event_id, organizer=request.user)
    regs  = EventRegistration.objects.filter(
        event=event, student__username__icontains=query
    ).select_related('student')
    return JsonResponse([r.student.username for r in regs], safe=False)


# ===========================================================
#  MY CERTIFICATES (Student)
# ===========================================================

@login_required
def my_certificates(request):
    certificates = StudentCertificate.objects.filter(
        student=request.user
    ).select_related('event')
    return render(request, "my_certificates.html", {"certificates": certificates})


# ===========================================================
#  PREVIOUS EVENTS
# ===========================================================

def prev_events(request):
    today  = timezone.now().date()
    events = Event.objects.filter(
        date__lt=today, status="APPROVED"
    ).order_by("-date").select_related('venue')
    registered_students = EventRegistration.objects.filter(event__in=events).count()
    return render(request, "previous_events.html", {
        "events":        events,
        "registrations": registered_students,
    })

@login_required
def submit_feedback(request, event_id):
    event = get_object_or_404(Event, id=event_id)
    reg   = get_object_or_404(EventRegistration,
                              student=request.user, event=event, verified=True)
    already = EventFeedback.objects.filter(student=request.user, event=event).exists()
    if already:
        messages.info(request, "You have already submitted feedback.")
        return redirect("my_registrations")
    if request.method == "POST":
        rating  = request.POST.get("rating")
        comment = request.POST.get("comment", "").strip()
        if rating and rating.isdigit() and 1 <= int(rating) <= 5:
            EventFeedback.objects.create(
                student=request.user, event=event,
                rating=int(rating), comment=comment or None
            )
            messages.success(request, "Thanks for your feedback!")
        else:
            messages.error(request, "Please select a rating between 1 and 5.")
        return redirect("my_registrations")
    return render(request, "submit_feedback.html", {"event": event})

@login_required
def event_feedback_summary(request, event_id):
    user = request.user
    event = get_object_or_404(Event, id=event_id)
    # Fetch Feedbacks
    feedbacks = EventFeedback.objects.filter(
        event=event
    ).select_related('student')
    total_reviews = feedbacks.count()
    avg = feedbacks.aggregate(
        avg=models.Avg('rating')
    )['avg']

    # Rating Distribution
    rating_distribution = {
        5: 0,
        4: 0,
        3: 0,
        2: 0,
        1: 0,
    }

    for fb in feedbacks:
        rating_distribution[fb.rating] += 1

    # Chart Data
    chart_data = []

    for rating, count in rating_distribution.items():
        percentage = (
            (count / total_reviews) * 100
            if total_reviews > 0 else 0
        )

        chart_data.append({
            "rating": rating,
            "count": count,
            "percentage": percentage,
        })

    return render(request, "event_feedback_summary.html", {
        
        "event": event,
        "feedbacks": feedbacks,
        "avg_rating": round(avg, 1) if avg else 0,
        "chart_data": chart_data,
        "total_reviews": total_reviews,
    })


@login_required
def broadcast_event_message(request, event_id):
    event = get_object_or_404(Event, id=event_id, organizer=request.user)
    if request.method == 'POST':
        msg = request.POST.get('message', '').strip()
        if msg:
            regs = EventRegistration.objects.filter(
                event=event
            ).select_related('student')
            for reg in regs:
                create_notification(
                    role='STUDENT', user=reg.student,
                    title=f"Update: {event.title}",
                    message=msg, event=event
                )
            messages.success(request, f"Message sent to {regs.count()} students.")
        return redirect('teacher_my_events')
    return render(request, 'broadcast_message.html', {'event': event})

# ============================================================
#  ADMIN: LOGISTICS & PASS MANAGEMENT OVERVIEW
# ============================================================
@login_required
def admin_logistics_overview(request):
    """Admin sees all approved/published events with pass status and assignment."""
    if not request.user.groups.filter(name='Admin').exists():
        messages.error(request, "Access denied. Admin privileges required.")
        return redirect('admin_dashboard')

    events = Event.objects.filter(
        status__in=['APPROVED']
    ).select_related(
        'venue', 'organizer', 'logistics_assignment__assigned_teacher'
    ).prefetch_related(
        'eventregistration_set'
    ).order_by('-date')

    event_data = []
    for event in events:
        regs = event.eventregistration_set.all()
        total = regs.count()
        active = regs.filter(pass_active=True).count()
        verified = regs.filter(verified=True).count()
        has_explicit_assignment = False
        try:
            assignment = event.logistics_assignment
            assigned_teacher = assignment.assigned_teacher
            has_explicit_assignment = True
        except Exception: # Handles DoesNotExist across one-to-one relations safely
            assigned_teacher = event.organizer

        event_data.append({
            'event': event,
            'total_registered': total,
            'passes_active': active,
            'passes_verified': verified,
            'assigned_teacher': assigned_teacher,
            'has_explicit_assignment': has_explicit_assignment,
            'is_published': event.is_published,
        })

    teachers = User.objects.filter(groups__name='Teacher').select_related('profile')

    return render(request, 'admin_logistics.html', {
        'event_data': event_data,
        'teachers': teachers,
    })


@login_required
def teacher_logistics_dashboard(request):
    """
    Fetches events where the logged-in teacher is either:
    1. The original explicit event organizer.
    2. Or assigned via the EventLogisticsAssignment model relationship.
    """
    if not request.user.groups.filter(name='Teacher').exists():
        messages.error(request, "Access denied. Teacher privileges required.")
        return redirect('main_page')

    # FIX: We query using the lower-case related name of the Assignment model table
    managed_events = Event.objects.filter(
        Q(organizer=request.user) | Q(logistics_assignment__assigned_teacher=request.user),
        status='APPROVED'
    ).select_related(
        'venue', 'organizer', 'logistics_assignment__assigned_teacher'
    ).distinct().order_by('-date')

    event_data_list = []
    for event in managed_events:
        regs = event.eventregistration_set.all()
        
        # Check if they are here because of a delegated assignment or structural ownership
        is_explicit_backup = False
        try:
            # Safe check because logistics_assignment is a OneToOneField relationship
            if hasattr(event, 'logistics_assignment') and event.logistics_assignment.assigned_teacher == request.user:
                if event.organizer != request.user:
                    is_explicit_backup = True
        except Exception:
            pass

        event_data_list.append({
            'event': event,
            'total_registered': regs.count(),
            'passes_active': regs.filter(pass_active=True).count(),
            'passes_verified': regs.filter(verified=True).count(),
            'is_explicit_backup': is_explicit_backup,
        })

    return render(request, 'teacher_logistics.html', {
        'managed_events': event_data_list
    })

# ============================================================
#  ADMIN: ASSIGN TEACHER TO LOGISTICS
# ============================================================
@login_required
@require_POST
def assign_logistics_teacher(request, event_id):
    """Admin assigns a teacher to manage passes/scanning for an event."""
    if not request.user.groups.filter(name='Admin').exists():
        messages.error(request, "Unauthorized execution access.")
        return redirect('admin_dashboard')

    event = get_object_or_404(Event, id=event_id)
    teacher_id = request.POST.get('teacher_id')
    notes = request.POST.get('notes', '')
    if teacher_id:
        teacher = get_object_or_404(User, id=teacher_id)
        if not teacher.groups.filter(name__in=['Teacher', 'Admin']).exists():
            messages.error(request, f"{teacher.get_full_name()} is not a Teacher or Admin and cannot be assigned.")
            return redirect('admin_logistics_overview')
        # Create or update assignment
        assignment, created = EventLogisticsAssignment.objects.update_or_create(
        event=event,
        defaults={
            'assigned_teacher': teacher,
            'assigned_by': request.user,
            'notes': notes,
            'assigned_at': timezone.now()
            }
        )
        create_notification(
            role='TEACHER',
            user=teacher,
            title='Logistics Assignment',
            message=f"You have been assigned as Logistics Manager for '{event.title}' on {event.date}. You can now publish passes and scan QR codes for this event.",
            event=event,
        )
        if event.organizer != teacher and event.organizer != request.user:
            create_notification(
                role='TEACHER',
                user=event.organizer,
                title='Logistics Manager Assigned to Your Event',
                message=f"Admin has assigned {teacher.get_full_name()} to manage passes and scanning for '{event.title}'. You remain the event organizer.",
                event=event,
            )

        log_action(
            request.user, 'LOGISTICS_ASSIGNED',
            event=event, target_user=teacher,
            notes=f"Assigned {teacher.get_full_name()} as logistics manager. Notes: {notes}",
            request=request
        )

        messages.success(request, f"Successfully delegated logistics execution for '{event.title}' to {teacher.get_full_name() or teacher.username}.")
    else:
        messages.error(request, "Invalid deployment parameter data.")

    return redirect('admin_logistics_overview')



# ===========================================================
#  MY WAITLIST (Student)
# ===========================================================
@login_required
def my_waitlist(request):
    waitlist_entries = WaitlistEntry.objects.filter(
        student=request.user
    ).select_related('event', 'event__venue').order_by('joined_at')

    # Annotate position for each entry
    entries_with_position = []
    for entry in waitlist_entries:
        position = WaitlistEntry.objects.filter(
            event=entry.event,
            joined_at__lt=entry.joined_at
        ).count() + 1
        entries_with_position.append({
            'entry': entry,
            'position': position,
            'total_waiting': WaitlistEntry.objects.filter(event=entry.event).count()
        })

    return render(request, 'my_waitlist.html', {
        'entries_with_position': entries_with_position
    })


@login_required
def leave_waitlist(request, event_id):
    entry = get_object_or_404(WaitlistEntry, student=request.user, event_id=event_id)
    if request.method == 'POST':
        entry.delete()
        messages.success(request, "You've been removed from the waitlist.")
    return redirect('my_waitlist')

@login_required
def unpublish_passes(request, event_id):
    if not request.user.groups.filter(name='Admin').exists():
        messages.error(request, "Admin only.")
        return redirect('admin_dashboard')
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'POST':
        EventRegistration.objects.filter(event=event).update(pass_active=False)
        event.is_published = False
        event.save(update_fields=['is_published'])
        log_action(request.user, 'PASS_DEACTIVATED', event=event,
           notes=f'All passes deactivated for: {event.title}', request=request)
        messages.success(request, f"Passes for '{event.title}' have been deactivated.")
    return redirect('admin_logistics_overview')